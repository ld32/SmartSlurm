#!/usr/bin/env python3
import sys
import os.path
import openpyxl

if len(sys.argv) != 2:
    print("Usage: " + os.path.basename(__file__) + "  <sampleList.xlsx, required.>")
    print("For example, please check online for Excel table format:")
    print("https://wiki.rc.hms.harvard.edu/display/O2/Build+Folder+Structures+From+Sample+Sheet+for+Rcbio+NGS+Workflows")
    exit()

sampleSheet = sys.argv[1]

if not os.path.isfile(sampleSheet):
    print("Sample sheet not exist: " + sampleSheet)
    exit()

wb = openpyxl.load_workbook(sampleSheet)
sh = wb.active

for i, (group, sample, library, lane, path, read1, read2, empty) in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
    group = str(group)
    print(group + '/' + sample + '/' + library + '/' + lane + '/' + path +'/' + read1 + "/" + read2)

    if ' ' in group or ' ' in sample or ' ' in library or ' ' in lane or ' ' in path or ' ' in read1:
        print(f"Row {i}: One of the column values contains a space.")
        exit()

    group = "smartSlurmInputs/" + group 
    if os.path.exists(group):
        print("Folder exists. please remove or rename it: " + group)
        exit()

    read1 = os.path.expanduser(os.path.join(path, read1))    
    if not os.path.isfile(read1):
        print(f"Row {i}: read1 file not exist: {read1}")
        exit()
    read2 = os.path.expanduser(os.path.join(path, read2))    
    if not os.path.isfile(read1):
        print(f"Row {i}: read1 file not exist: {read2}")

for i, (group, sample, library, lane, path, read1, read2, empty) in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
     
    group = 'smartSlurmInputs/' + str(group)    
    read1 = os.path.expanduser(os.path.join(path, read1))
    
    if not os.path.exists(group + "/" + sample + "/"):
        os.makedirs(group + "/" + sample + "/")
    if read1.endswith(".gz"):
        os.symlink(read1, group + "/" + sample + "/" + library + "_" + lane + "_1.fq.gz")
    elif read1.endswith(".bz2"):
        os.symlink(read1, group + "/" + sample + "/" + library + "_" + lane + "_1.fq.bz2")
    else:
        os.symlink(read1, group + "/" + sample + "/" + library + "_" + lane + "_1.fq")
    
    if read2:
        read2 = os.path.expanduser(os.path.join(path, read2))

        if not os.path.isfile(read2):
            print(f"Row {i}: read2 file not exist: {read2}")
            exit()
        if read2.endswith(".gz"):
            os.symlink(read2, group + "/" + sample + "/" + library + "_" + lane + "_2.fq.gz")
        elif read2.endswith(".bz2"):
            os.symlink(read2, group + "/" + sample + "/" + library + "_" + lane + "_2.fq.bz2")
        else:
            os.symlink(read2, group + "/" + sample + "/" + library + "_" + lane + "_2.fq")

print("Done")
