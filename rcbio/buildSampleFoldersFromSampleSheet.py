#!/usr/bin/env python3

import openpyxl, sys

import os.path

if len(sys.argv) != 2:
        print("Usage: " + os.path.basename(__file__) + "  <sampleList.xlsx, required.>")
        print("For example, please check online for Excel table format:")
        print("https://wiki.rc.hms.harvard.edu/display/O2/Build+Folder+Structures+From+Sample+Sheet+for+Rcbio+NGS+Workflows")
        sys.exit()

sampleSheet=sys.argv[1]

if not os.path.isfile(sampleSheet):
        print("Sample sheet not exist: " + sampleSheet)
        sys.exit()

if os.path.exists("group1"):
    print("Folder exists. please remove or rename it: group1")
    sys.exit()

if os.path.exists("group2"):
    print("Folder exists. please remove or rename it: group2")
    sys.exit()
    
if os.path.exists("group3"):
    print("Folder exists. please remove or rename it: group3")
    sys.exit()    

wb = openpyxl.load_workbook(sampleSheet)
sh = wb.active

for i in range(1, sh.max_row + 1):
    if i==1: # ignore header 
        continue
    
    print(" checking row: " + str(i))
    group = sh.cell(row=i, column=1).value
               
    if type(group) is float:        
        group=str(group).split(".")[0]
    
    if ' ' in str(group): 
        print("Group name can not have space: '%s'" % group)
        sys.exit()
        
    sample = sh.cell(row=i, column=2).value
    if type(sample) is float:        
        sample=str(sample).split(".")[0]
 
    if ' ' in sample: 
        print("Sample name can not have space: '%s'" % sample)
        sys.exit()

    library = sh.cell(row=i, column=3).value
    if type(library) is float:        
        library=str(library).split(".")[0]
    
    if ' ' in library: 
        print("Library name can not have space: '%s'" % library)
        sys.exit()
    
    lane = sh.cell(row=i, column=4).value
    if type(lane) is float:        
        lane=str(lane).split(".")[0]
    
    if ' ' in lane: 
        print("Lane name can not have space: '%s'" % lane)
        sys.exit()
    
    path = sh.cell(row=i, column=5).value
    if type(path) is float:        
        path=str(path).split(".")[0]
    
    if ' ' in path: 
        print("Path name can not have space: '%s'" % path)
        sys.exit()
        
    read1 = os.path.expanduser(path + "/" + sh.cell(row=i, column=6).value)
    
    if ' ' in read1: 
        print("Read1 name can not have space: '%s'" % read1)
        sys.exit()
    
    if not os.path.isfile(read1): 
        print("row " + str(i) + ": read1 file not exist: " + read1)
        sys.exit()
    if sh.cell(row=i, column=7).value != "" and sh.cell(row=i, column=7).value is not None:  # if we have read2
        read2 = os.path.expanduser(path + "/" + sh.cell(row=i, column=7).value)
        if ' ' in read2: 
            print("Read2 name can not have space: '%s'" % read2)
            sys.exit()
        
        if not os.path.isfile(read2): 
            print("row " + str(i) + ": read2 file not exist: " + read2)
            sys.exit()
            
for i in range(1, sh.max_row + 1):
    if i==1: # ignore header 
        continue
    
    print("working on row: " + str(i))
    group = sh.cell(row=i, column=1).value
               
    if type(group) is float:        
        group=str(group).split(".")[0]
    else:
        group = str(group)  # Ensure group is always a string
    
    sample = sh.cell(row=i, column=2).value
    if type(sample) is float:        
        sample=str(sample).split(".")[0]
    else:
        sample = str(sample)  # Ensure sample is always a string
 
    library = sh.cell(row=i, column=3).value
    if type(library) is float:        
        library=str(library).split(".")[0]
    else:
        library = str(library)  # Ensure library is always a string
    
    lane = sh.cell(row=i, column=4).value
    if type(lane) is float:        
        lane=str(lane).split(".")[0]
    else:
        lane = str(lane)  # Ensure lane is always a string
    
    path = sh.cell(row=i, column=5).value
    if type(path) is float:        
        path=str(path).split(".")[0]
    else:
        path = str(path)  # Ensure path is always a string
    
    read1 = os.path.expanduser(path + "/" + sh.cell(row=i, column=6).value)
    
    if not os.path.exists("group" + group + "/" + sample + "/" ): 
        os.makedirs("group" + group + "/" + sample + "/" )
    if read1.endswith(".gz"):  
        os.symlink(read1, "group" + group + "/" + sample + "/" + library + "_" + lane +"_1.fq.gz" )
    elif read1.endswith(".bz2"):
        os.symlink(read1, "group" + group + "/" + sample + "/" + library + "_" + lane + "_1.fq.bz2" )
    else:
        os.symlink(read1, "group" + group + "/" + sample + "/" + library + "_" + lane + "_1.fq" )
    if sh.cell(row=i, column=7).value != "" and sh.cell(row=i, column=7).value is not None:  # if we have read2
        read2 = os.path.expanduser(path + "/" + sh.cell(row=i, column=7).value)
       
        if read2.endswith(".gz"):  
            os.symlink(read2, "group" + group + "/" + sample + "/" + library + "_" + lane + "_2.fq.gz" )
        elif read2.endswith(".bz2"):
            os.symlink(read2, "group" + group + "/" + sample + "/" + library + "_" + lane + "_2.fq.bz2" )
        else:
            os.symlink(read2, "group" + group + "/" + sample + "/" + library + "_" + lane + "_2.fq" )
                  
if not os.path.exists("group2"):
    print("Warning: Folder group2 is not created. Usually you should have at least two groups for NGS analysis.")
    sys.exit()

print("Done")


